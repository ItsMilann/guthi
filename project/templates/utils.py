import nepali_datetime
from django.template.loader import render_to_string
from django.core.files.base import ContentFile
from django.core.files.storage import FileSystemStorage
from weasyprint import HTML

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

from utilities.utils import get_nepali_number
from branches.models import Branch
from branches.models import FiscalYear


def get_excel_report(data, date_string):
    nagarpalika = Branch.objects.nagarpalika().latest("id")

    wb = Workbook()
    ws = wb.active
    ws.title = "Patrachar Report"

    border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    fill = PatternFill(
        fill_type=None,
        start_color='00FFCC00',
        end_color='00FFCC00'
    )

    alignment = Alignment(
        horizontal='left',
        vertical='center',
        shrink_to_fit=True
    )

    center_alignment = Alignment(
        horizontal='center',
        vertical='center',
        shrink_to_fit=True
    )

    for i in range(3):
        ws.row_dimensions[i].height = 25

    ws.merge_cells('A1:D1')
    ws.merge_cells('A2:D2')
    ws.merge_cells('A3:D3')
    ws.merge_cells('A4:D4')
    ws.merge_cells('A5:D5')

    h1 = ws['A1']
    h2 = ws['A2']
    h3 = ws['A3']
    h4 = ws['A4']
    h5 = ws['A5']
    h1.font, h2.font, h3.font, h4.font, h5.font = [
        Font(bold=True, size=12, color='FF0000'),
        Font(bold=True, size=12, color='FF0000'),
        Font(bold=True, size=12, color='FF0000'),
        Font(bold=True, size=12, color='FF0000'),
        Font(bold=True, size=12, color='FF0000')
    ]
    h1.value = "सिफारिस प्रणाली"
    h2.value = nagarpalika.name_np or nagarpalika.name_en
    type_ = "गाउँ" if "Rural" in nagarpalika.name_en else "नगर"
    h3.value = f"{type_} कार्यपालिकाको कार्यालय"
    h4.value = nagarpalika.address_np or nagarpalika.address_en
    h5.value = date_string

    h1.alignment, h2.alignment, h3.alignment, h4.alignment, h5.alignment = [
        center_alignment, center_alignment, center_alignment, center_alignment, center_alignment
    ]

    start_row = 7
    others_width = 10
    row = 12

    ws.column_dimensions[get_column_letter(2)].width = 35
    ws.column_dimensions[get_column_letter(1)].width = 5
    ws.column_dimensions[get_column_letter(3)].width = 25
    ws.column_dimensions[get_column_letter(4)].width = 20
    ws.column_dimensions[get_column_letter(5)].width = 20

    first_table_data = {
        "templates": data["templates"],
        "papers": data["papers"],
        "applicant": data["applicant"],
        "worth": data["worth"],
    }
    second_table_data = data["template_report"]

    ws.cell(start_row, 1, "क्र.स")
    ws.cell(start_row, 1).font = Font(bold=True, size=12, color='FFFFFF')
    ws.cell(start_row, 1).alignment = alignment
    ws.cell(start_row, 1).border = border
    ws.cell(start_row, 1).fill = PatternFill(
        fill_type='solid', start_color='FF0000', end_color='FF0000')

    ws.cell(start_row, 2, "Category Name")
    ws.cell(start_row, 2).font = Font(bold=True, size=12, color='FFFFFF')
    ws.cell(start_row, 2).alignment = alignment
    ws.cell(start_row, 2).border = border
    ws.cell(start_row, 2).fill = PatternFill(
        fill_type='solid', start_color='FF0000', end_color='FF0000')

    ws.cell(start_row, 3, "Issued Applications")
    ws.cell(start_row, 3).font = Font(bold=True, size=12, color='FFFFFF')
    ws.cell(start_row, 3).alignment = alignment
    ws.cell(start_row, 3).border = border
    ws.cell(start_row, 3).fill = PatternFill(
        fill_type='solid', start_color='FF0000', end_color='FF0000')
    ws.cell(start_row, 4, "Total Amount")
    ws.cell(start_row, 4).font = Font(bold=True, size=12, color='FFFFFF')
    ws.cell(start_row, 4).alignment = alignment
    ws.cell(start_row, 4).border = border
    ws.cell(start_row, 4).fill = PatternFill(
        fill_type='solid', start_color='FF0000', end_color='FF0000')

    row = start_row + 1
    data = second_table_data
    for i in range(len(data)):
        ws.cell(i+row, 1, i+1)
        ws.cell(i+row, 1).alignment = alignment
        ws.cell(i+row, 1).border = border
        ws.cell(i+row, 1).fill = fill

        ws.cell(i+row, 2, data[i]["title"])
        ws.cell(i+row, 2).alignment = alignment
        ws.cell(i+row, 2).border = border
        ws.cell(i+row, 2).fill = fill

        ws.cell(i+row, 3, data[i]["total_papers"])
        ws.cell(i+row, 3).alignment = alignment
        ws.cell(i+row, 3).border = border
        ws.cell(i+row, 3).fill = fill

        ws.cell(i+row, 4, data[i]["total_amount"])
        ws.cell(i+row, 4).alignment = alignment
        ws.cell(i+row, 4).border = border
        ws.cell(i+row, 4).fill = fill

    ws.cell(start_row+len(data)+1, 1, "")
    ws.cell(start_row+len(data)+1, 1).alignment = alignment
    ws.cell(start_row+len(data)+1, 1).border = border
    ws.cell(start_row+len(data)+1, 1).fill = fill

    ws.cell(start_row+len(data)+1, 2, "Total")
    ws.cell(start_row+len(data)+1, 2).alignment = alignment
    ws.cell(start_row+len(data)+1, 2).border = border
    ws.cell(start_row+len(data)+1, 2).fill = fill

    ws.cell(start_row+len(data)+1, 3, first_table_data["papers"])
    ws.cell(start_row+len(data)+1, 3).alignment = alignment
    ws.cell(start_row+len(data)+1, 3).border = border
    ws.cell(start_row+len(data)+1, 3).fill = fill

    ws.cell(start_row+len(data)+1, 4, first_table_data["worth"])
    ws.cell(start_row+len(data)+1, 4).alignment = alignment
    ws.cell(start_row+len(data)+1, 4).border = border
    ws.cell(start_row+len(data)+1, 4).fill = fill

    path = f"media/sifarish-excel-report.xlsx"
    wb.save(path)
    return path


def get_pdf_report(data, listing_data, date_string, ward_number, file_name=None):
    fiscal_year = FiscalYear.active().title
    total_applicants = data.get("templates", 0)
    total_sifarish = data.get("papers", 0)
    total_papers = data.get("applicant", 0)
    total_amount = data.get("worth", 0)
    nagarpalika = Branch.objects.nagarpalika().latest("id")
    html = render_to_string('pdf_template.html', {
        "second_table_data": listing_data,
        "date_string": date_string,
        "fiscal_year": fiscal_year,
        "ward_number": list(ward_number),
        "total_applicants": get_nepali_number(total_applicants),
        "total_sifarish": get_nepali_number(total_sifarish),
        "total_papers": get_nepali_number(total_papers),
        "total_amount": get_nepali_number(total_amount),
        "nagarpalika": nagarpalika
    })
    pdf = HTML(string=html).write_pdf()
    c = ContentFile(pdf)
    fs = FileSystemStorage()
    if not file_name:
        file_name = "sifarish-pdf-report"
    filename = fs.save(f"{file_name}.pdf", c)
    uploaded_file_url = fs.url(filename)
    return uploaded_file_url


def generate_issue_number(paper):
    count = paper.__class__.objects.filter(template = paper.template).count()
    date = nepali_datetime.date.today()
    month, day = date.month, date.day
    fiscal_year = FiscalYear.active()
    fy_text = "".join([t for t in fiscal_year.title if t.isalnum()])
    return f"{fy_text}-{month:02d}-{day:02d}-{paper.id}-{count}"
